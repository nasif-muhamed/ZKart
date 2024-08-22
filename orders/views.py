from django.shortcuts import render, HttpResponse, redirect
from . models import *
from users . models import *
from users . views import *
from products . models import *
from products . views import *
from django.db.models import F, ExpressionWrapper, DateTimeField
from django.db.models.functions import Lower
# from django.core.paginator import Paginator
# from django.db.models import Q
from django.http import JsonResponse
import json
from django.utils import timezone
from django.contrib.auth.decorators import login_required


@login_required(login_url='/')
def cart_view(request):
    
    customer = request.user.account
    order, created = Order.objects.get_or_create(customer=customer, status='cart')
    orders = order.order_items.all()
    sub_total = order.get_cart_total()
    total_discount = order.get_cart_discount()
    delivery_charge = 150 if sub_total < 2000 else 0
    coupon_discount = request.session['coupon_discount'] if 'coupon_discount' in request.session else 0
    coupon_applicable = Coupon.objects.filter(is_expired = False)

    if request.method == 'POST' and 'cart_submit' in request.POST:
        request.session['checkout_access'] = True
        return redirect(checkout_page)

    if request.method == 'POST':
        coupon_code = request.POST.get('coupon_code')
        if 'apply_coupon' in request.POST:
            try:
                coupon = coupon_applicable.get(coupon_code = coupon_code)
                if sub_total > coupon.minimum_amount:
                    request.session['applied_coupon'] = coupon_code

                    if coupon.type == 'amount':
                        request.session['coupon_discount'] = coupon.discount
                    else:
                        request.session['coupon_discount'] = round((sub_total * coupon.discount) / 100)

                    messages.success(request, 'Coupon applied successfully.')

                else:
                    messages.error(request, 'Coupon code not applicable for this order.')

            except Coupon.DoesNotExist:
                messages.error(request, 'Invalid coupon code.')
        
        elif 'remove_coupon' in request.POST:
            if 'applied_coupon' in request.session:
                del request.session['applied_coupon']

            if 'coupon_discount' in request.session:
                del request.session['coupon_discount']

            messages.info(request, 'Coupon removed successfully.')

        
        return redirect(cart_view)

    total_amount = sub_total - coupon_discount
    
    context = {
        'orders' : orders,
        'sub_total' : sub_total,
        'total_discount' : total_discount,
        'delivery_charge' : delivery_charge,
        'total_amount' : total_amount,
        'coupon_applicable' : coupon_applicable,
    }
    return render(request,'cart.html', context)


@login_required(login_url='/')
def add_to_cart(request):
    variant_id = int(request.POST.get('variant_id'))
    quantity = int(request.POST.get('quantity'))
    add_success = False

    variant = ProductVariant.objects.get(id = variant_id)
    customer = request.user.account

    max_quantity = variant.product.max_purchase_qty
    if quantity > max_quantity:
        messages.error(request, f'A person can only purchase this product upto {max_quantity} quantity')

    elif quantity > variant.quantity:
        messages.error(request, 'Insufficient Stock')

    else:
        order, order_created = Order.objects.get_or_create(customer=customer, status='cart')
        order_item, item_created = OrderItem.objects.get_or_create(product_variant = variant, order = order)
        order_item.quantity
        
        if item_created:
            order_item.quantity = quantity
            messages.success(request, 'Product added to cart successfully')
            order_item.save()
            add_success = True

        else:
            if (order_item.quantity + quantity) > max_quantity:
                messages.error(request, f'A person can only purchase this product upto {max_quantity} quantities')

            elif (order_item.quantity + quantity) > variant.quantity:
                messages.error(request, 'Insufficient Stock')

            else:
                order_item.quantity += quantity
                messages.success(request, 'Product on cart is updated')
                order_item.save()
                add_success = True

        if add_success:
            if 'coupon_discount' in request.session:
                del request.session['coupon_discount']
                del request.session['applied_coupon']
                # order.coupon = None
                # order.save()
                messages.info(request, 'Coupon in cart has been reset due to change in order.')

    return redirect(product_details, variant.product.id) 


@login_required(login_url='/')
def update_cart(request):
    data = json.loads(request.body)
    item_id = data['item_id']
    action = data['action']
    order_item = OrderItem.objects.get(id=item_id)

    if action == 'add':
        order_item.quantity = (order_item.quantity + 1)
    
    elif action == 'remove':
        order_item.quantity = (order_item.quantity - 1)

    order_item.save()

    if 'coupon_discount' in request.session:
        del request.session['coupon_discount']
        del request.session['applied_coupon']
        # order.coupon = None
        # order.save()
        messages.info(request, 'Coupon has been reset due to change in order.')

    return JsonResponse({'success': True, })     


@login_required(login_url='/')
def delete_cart_item(request):
    
    if 'coupon_discount' in request.session:
        del request.session['coupon_discount']
        del request.session['applied_coupon']
        # order.coupon = None
        # order.save()
        messages.info(request, 'Coupon has been reset due to item removal.')
    
    data = json.loads(request.body)
    item_id = data['item_id']
    item = OrderItem.objects.get(id = int(item_id))
    item.delete()
    return JsonResponse({'success': True, })     

import random
import string

def generate_unique_order_id(customer_id):
    current_time = datetime.now().strftime('%m%d%H%M%S') 
    customer_id = str(customer_id)
    random_chars = ''.join(random.choices(string.ascii_uppercase + string.digits, k=2))
    order_id = f"{current_time}{customer_id}{random_chars}"
    order_id = order_id
    
    while Order.objects.filter(order_identifier=order_id).exists():
        random_chars = ''.join(random.choices(string.ascii_uppercase + string.digits, k=2))
        order_id = f"{current_time}{customer_id}{random_chars}"
        order_id = order_id

    return order_id

from django.http import JsonResponse
from geopy.geocoders import Nominatim
from geopy.distance import geodesic
    
def get_delivery_charge(address_id):
    try:
        SELLER_HUB_COORDINATES = (9.9312, 76.2673)  # Kochi
        address = Address.objects.get(id=address_id)
        address_latitude = address.latitude
        address_longitude = address.longitude
        geolocator = Nominatim(user_agent="zkart")

        if address_latitude is not None and address_longitude is not None:
            user_coordinates = (address_latitude, address_longitude)

        else:
            location = geolocator.geocode(f'{address.city}, {address.pin_code}')

            if location:
                user_coordinates = (location.latitude, location.longitude)
            
            else:
                delivery_charge = 'None'
                return delivery_charge
            
        distance_km = geodesic(SELLER_HUB_COORDINATES, user_coordinates).km

        if distance_km <= 100:
            delivery_charge = 'Free' 
        
        else:
            distance_km -= 100
            delivery_charge = math.ceil(distance_km/50) * 20

        return delivery_charge
    
    except:
        return 'None'

def get_delivery_charge_for_checkout(request):
    if request.method == 'POST':
        address_id = request.POST.get('address_id')
        delivery_charge = get_delivery_charge(address_id)
        return JsonResponse({'delivery_charge': delivery_charge})


def order_address_creator(og_address):
    order_address = OrderAddress()
    order_address.name = og_address.name
    order_address.mobile = og_address.mobile
    order_address.address_line1 = og_address.address_line1
    order_address.address_line2 = og_address.address_line2
    order_address.city = og_address.city
    order_address.state = og_address.state
    order_address.pin_code = og_address.pin_code
    order_address.latitude = og_address.latitude
    order_address.longitude = og_address.longitude
    order_address.country = og_address.country
    order_address.save()
    return order_address


@login_required(login_url='/')
def checkout_page(request):
    if request.method == 'GET' and not request.session.get('checkout_access'):
        return redirect(cart_view)
 
    request.session['checkout_access'] = False

    customer = request.user.account
    if not customer.is_completed():
        messages.error(request, 'Complete your profile prior to Checkout')
        return redirect(cart_view)

    addresses = customer.user_addresses.all()
    order, created = Order.objects.get_or_create(customer=customer, status='cart')
    orders = order.order_items.all()

    if orders:
        for order_item in orders:
            if order_item.quantity > order_item.product_variant.quantity:
                messages.error(request, f'{order_item.product_variant.product.title} is out of stock. Please remove it from the cart or adjust the stock')
                return redirect(cart_view)
        sub_total = order.get_cart_total()
        delivery_charge = 150 if sub_total < 2000 else 0
        coupon_discount = request.session['coupon_discount'] if 'coupon_discount' in request.session else 0
        total_amount = sub_total - coupon_discount

        if request.method == 'POST':
            selected_address_id = request.POST.get('selectedAddress')
            selected_payment_method = request.POST.get('selectedPayment')
            special_instructions = request.POST.get('special_instructions')
            payment_id = request.POST.get('payment_id')
            payment_failed = request.POST.get('payment_failed')

            if "cod_button" in request.POST:
                selected_payment_method = 'cod'
                if total_amount > 1000:
                    messages.error(request, 'Order above Rs 1000 is not eligible for COD')
                    return redirect(cart_view)


            elif "wallet_button" in request.POST:
                selected_payment_method = 'wallet'
                customer_wallet = customer.wallet
                if customer_wallet.balance < total_amount:
                    messages.error(request, 'Not enough balance in your Wallet')
                    return redirect(cart_view)

            try:
                if selected_address_id and selected_payment_method:
                    og_address = Address.objects.get(id=int(selected_address_id))
                    delivery_charge = get_delivery_charge(og_address.id)
                    if delivery_charge == 'Free':
                        delivery_charge = 0
                    elif delivery_charge == 'None':
                        messages.error(request, 'Invalid Address')
                        return redirect(cart_view)
                    
                    order_address = order_address_creator(og_address)
                    order.address = order_address
                    order.payment_method = selected_payment_method
                    order.payment_id = payment_id
                    if special_instructions:
                        order.special_instructions = special_instructions

                    if selected_payment_method == 'razorpay' and payment_failed == 'true':
                        order.status = 'pending'
                    else:
                        order.status = 'placed'
                        order.order_date = timezone.now()

                    order.delivery_charge = delivery_charge

                    if selected_payment_method == 'cod' or selected_payment_method == 'wallet':
                        order.order_identifier = generate_unique_order_id(customer.id)
                        
                    order.save()

                    for order_item in orders:
                        if selected_payment_method == 'razorpay' and payment_failed == 'true':
                            order_item.status = 'pending'
                            order_item.payment_status = 'failed'

                        else:
                            order_item.status = 'in_progress'
                            order_item.order_date = timezone.now()                    
                            order_item.product_variant.quantity -= order_item.quantity
                            if selected_payment_method == 'razorpay' or selected_payment_method == 'wallet':
                                order_item.payment_status = 'success'
                            
                        order_item.selling_price = order_item.product_variant.product.product_selling_price()
                        order_item.original_price = order_item.product_variant.product.original_price

                        order_item.save()
                        order_item.product_variant.save()
                        order_item.product_variant.product.save()

                    if 'coupon_discount' in request.session:
                        order.coupon = Coupon.objects.get(coupon_code= request.session['applied_coupon'])
                        order_total_basic = order.order_total_basic()
                        coupon_discounts = int(request.session['coupon_discount'])
                        for order_item in orders:
                            order_item.coupon_discount = round((order_item.selling_price * order_item.quantity / order_total_basic) * coupon_discounts)
                            order_item.save()

                        order.save()
                        del request.session['applied_coupon']
                        del request.session['coupon_discount']

                    if selected_payment_method == 'razorpay':
                        response_data = {
                            'status': 'success',
                            'message': f'Your Order ID: {order.id}',
                            'order_id': order.id,
                        }
                        return JsonResponse(response_data)

                    if selected_payment_method == 'wallet':
                        customer_wallet.balance -= total_amount
                        customer_wallet.save()

                    return redirect(order_success, order.id)

                else:
                    messages.error(request, 'Something went wrong')
                    return redirect(checkout_page)

            except:
                messages.error(request, 'Something went wrong')
                return redirect(checkout_page)


        context = {
            'orders' : orders,
            'addresses' : addresses,
            'sub_total' : sub_total,
            'delivery_charge' : delivery_charge,
            'total_amount' : total_amount,
            'coupon_discount_js' : coupon_discount,
        }
        return render(request,'checkout.html', context)
    
    return redirect(cart_view)


def retry_payment_stock_check(request):
    try:
        order_id = request.POST.get('order_id')
        order = Order.objects.get(id= int(order_id))
        order_identifier = order.order_identifier
        user = request.user
        address = order.address

        if order.status == 'pending' and order.customer.user == request.user:
            for item in order.order_items.all(): 
                if item.quantity > item.product_variant.quantity:
                    return JsonResponse({'status': 'error', 'message': 'Invalid JSON data', 'reason':'Out of stock Items found'}, status=400)

            response_data = {
                'status': 'success',
                'message': f'Your Order ID: {order_identifier}',
                'order_identifier': order_identifier,
                'name': address.name,
                'email': user.email,
                'contact': address.mobile,

            }
            return JsonResponse(response_data)
        else:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON data', 'reason':'Something went wrong'}, status=400)
    except:
        response_data = {
            'status': 'error',
            'message': f'something went wrong',
            'reason': 'Oops something went wrong',
        }
        return JsonResponse(response_data, status=405)


def retry_payment(request, order_id):
    try:
        payment_id = request.POST.get('payment_id')
        order_identifier = request.POST.get('order_identifier')
        customer = request.user.account
        order = Order.objects.get(customer=customer, order_identifier=order_identifier)
        orders = order.order_items.all()
        order.status = 'placed'
        order.order_date = timezone.now()
        order.payment_id = payment_id
        order.save()

        if orders:
            for order_item in orders:
                order_item.status = 'in_progress'
                order_item.order_date = timezone.now()                    
                order_item.product_variant.quantity -= order_item.quantity
                order_item.payment_status = 'success'
                order_item.save()
                order_item.product_variant.save()
                order_item.product_variant.product.save()

        response_data = {
            'status': 'success',
            'message': f'Your Order ID: {order.id}',
            'order_id': order.id,
        }
        return JsonResponse(response_data)
    
    except:
        response_data = {
            'status': 'error',
            'message': f'something went wrong',
            'order_id': order.id,
        }
        return JsonResponse(response_data)

from zkart.settings import RAZORPAY_API_KEY, RAZORPAY_API_SECRET_KEY
import razorpay

client = razorpay.Client(auth=(RAZORPAY_API_KEY, RAZORPAY_API_SECRET_KEY))
@login_required(login_url='/')
def razorpaycheck(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            selected_address = data.get('selectedAddress')
            address = Address.objects.get(id=int(selected_address))

            user = request.user
            order = Order.objects.get(customer=user.account, status = 'cart')
            sub_total = order.get_cart_total()
            delivery_charge = get_delivery_charge(address.id)
            if delivery_charge == 'Free':
                delivery_charge = 0
            elif  delivery_charge == 'None':
                return JsonResponse({'status': 'error', 'message': 'Invalid JSON data', 'reason':'Invalid Address Add/Select Another one'}, status=400)

            coupon_discount = request.session['coupon_discount'] if 'coupon_discount' in request.session else 0
            total_amount = sub_total + delivery_charge - coupon_discount

            order_currency = 'INR'
            payment_order = client.order.create(dict(amount=total_amount*100, currency=order_currency, payment_capture=1))
            payment_order_id = payment_order['id']

            order.order_identifier = payment_order_id
            order.save()

            response_data = {
                'status': 'success',
                'message': f'Selected address ID: {selected_address}',
                'name': address.name,
                'email': user.email,
                'contact': address.mobile,
                'api_key' : RAZORPAY_API_KEY,
                'order_id' : payment_order_id,
            }
            return JsonResponse(response_data)

        except json.JSONDecodeError:
                return JsonResponse({'status': 'error', 'message': 'Invalid JSON data'}, status=400)
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)


@login_required(login_url='/')
def order_success(request, order_id):
    order = Order.objects.get(id= int(order_id))
    if order.order_date:
        expected_arrival = order.order_date + timedelta(days=7)
    else:
        expected_arrival = None

    context = {
        'order' : order, 
        'expected_arrival' : expected_arrival
    }
    return render(request, 'order_success.html', context)


@login_required(login_url='/')
def user_orders_page(request):
    customer = request.user.account
    orders = Order.objects.filter(customer=customer).exclude(status='cart').annotate(
        expected_date=ExpressionWrapper(
            F('order_date') + timedelta(days=7),
            output_field=DateTimeField()
        )
    )

    context = {
        'orders' : orders,
        'account' : customer,
    }
    return render(request,'user_profile/manage_my_order.html', context)
    

@login_required(login_url='/')
def user_order_details(request, order_id):
    customer = request.user.account
    order = Order.objects.get(id=order_id)
    if order.order_date:
        expected_date = order.order_date + timedelta(days=7)
    else:
        expected_date = None

    if 'invoice_btn' in request.GET:
        response = render_to_pdf('user_profile/pdf_invoice.html', {'order':order})
        response['Content-Disposition'] = 'attachment; filename="tax_invoice.pdf"'
        return response

    if request.method == 'POST':
        if 'return_btn' in request.POST:
            order_item_id = request.POST.get('return_btn')
            order_item = OrderItem.objects.get(id=int(order_item_id))
            order_item.return_date = timezone.now()
            # checking return date exceeded or not
            if order_item.return_date > (order_item.order_date + timedelta(days=7)):
                messages.error(request, '7 days return Period is over')
                return redirect(user_order_details, order_id)

            order_item.status = 'return_request'
            order_item.save()

            email = 'muhdnasifk@gmail.com'
            mail_subject = f'Return Request of product delivered at {order_item.completed_date.strftime("%Y-%m-%d")}' 
            mail_message = render_to_string('emailer/return_request_email.html', {'order_item': order_item,
                                                                                    'user' : request.user.username,})
            emailer = EmailMessage(
                mail_subject, mail_message, to= [email] 
            )
            emailer.send()

        
        elif 'return_cancel_btn' in request.POST:
            order_item_id = request.POST.get('return_cancel_btn')
            order_item = OrderItem.objects.get(id=int(order_item_id))
            order_item.status = 'delivered'
            order_item.save()

        return redirect(user_order_details, order_id)

    context = {
        'order' : order,
        'account' : customer,
        'expected_date': expected_date,
        
    }
    
    return render(request,'user_profile/manage_order_details.html', context)


@login_required(login_url='/')
def user_order_cancel(request, order_item_id):
    order_item = OrderItem.objects.get(id=int(order_item_id))
    order_item.status = 'cancelled'
    order_item.completed_date = timezone.now()
    order_item.save()
    order_item.product_variant.quantity += order_item.quantity
    if order_item.payment_status == 'success':
        wallet = request.user.account.wallet
        wallet.deposit((order_item.selling_price*order_item.quantity)-order_item.coupon_discount)
        order_item.payment_status = 'wallet'
        wallet.save()
    else:
        order_item.payment_status = 'cancelled'

    order = order_item.order
    if order.is_completed():
        order.complete_date = timezone.now()
        order.status = 'completed'
        order.save()

    order_item.save()
    order_item.product_variant.product.save()
    order_item.product_variant.save()
    
    email = 'muhdnasifk@gmail.com'
    mail_subject = f'Order Item Cancelled'
    mail_message = render_to_string('emailer/order_cancel_email.html', {'order_item': order_item,
                                                                            'user' : request.user.username,})
    emailer = EmailMessage(
        mail_subject, mail_message, to= [email] 
    )
    emailer.send()

    return redirect(user_order_details, order.id)


# Admin side

@login_required(login_url='admin_login')
@user_passes_test(is_admin, login_url='/permission-denied/')
def order_management(request):
    orders = Order.objects.exclude(Q(status='cart') | Q(status='pending')).order_by('-order_date')
    search = request.GET.get('search')
    if search and search is not None:
        orders = orders.filter(Q(order_identifier__icontains=search) | Q(customer__user__username__icontains=search) | Q(order_items__status=search))

    context = {
        'orders' : orders,
        'search' : search,
    }
    return render(request,'admin_page/order_management/order_management.html', context)


@login_required(login_url='admin_login')
@user_passes_test(is_admin, login_url='/permission-denied/')
def return_items_management(request):
    return_request = OrderItem.objects.filter(status = 'return_request').order_by('completed_date')
    search = request.GET.get('search')
    if search and search is not None:
        return_request = return_request.filter(Q(order__order_identifier__icontains = search) | Q(product_variant__product__title__icontains=search))
    
    if request.method == 'POST':
        try:
            item_id = request.POST.get('item-id')
            item = return_request.get(id=int(item_id))

            if 'decline-btn' in request.POST and item.status == 'return_request':
                item.status = 'delivered'
                item.save()
                messages.success(request, 'Declined Return Request')

            elif 'accept-btn' in request.POST and item.status == 'return_request':
                item.status = 'returned'
                item.payment_status = 'wallet'
                wallet = item.order.customer.wallet
                wallet.deposit((item.selling_price*item.quantity)-item.coupon_discount) 
                item.product_variant.quantity += item.quantity
                item.save()
                item.product_variant.product.save()
                item.product_variant.save()
                messages.success(request, 'Accepted Return Request')
            
            return redirect(return_items_management)

        except:
            messages.error(request, 'Oops something went wrong')
            return redirect(return_items_management)

    context = {
        'return_request' : return_request,
        'search' : search,
    }
    return render(request,'admin_page/order_management/return_item.html', context)


@login_required(login_url='admin_login')
@user_passes_test(is_admin, login_url='/permission-denied/')
def order_item_management(request, order_id):
    order = Order.objects.get(id = int(order_id))
    order = order.order_items.all()
    context = {
        'order' : order,
    }
    return render(request,'admin_page/order_management/order_item_management.html', context)


@login_required(login_url='admin_login')
@user_passes_test(is_admin, login_url='/permission-denied/')
def update_order_status(request, order_item_id):
    if request.method == 'POST' and request.user.is_staff:
        order_item = OrderItem.objects.get(id=int(order_item_id))
        status = request.POST.get('status')
        order_item.status = status
        order = order_item.order
        
        if status == 'delivered' or status == 'cancelled':
            order_item.completed_date = timezone.now()
            order_item.save()

            if status == 'delivered':
                order_item.payment_status = 'success'

            else:
                order_item.product_variant.quantity += order_item.quantity
                if order_item.payment_status == 'success':
                    wallet = order_item.order.customer.wallet
                    wallet.deposit((order_item.selling_price*order_item.quantity)-order_item.coupon_discount)
                    order_item.payment_status = 'wallet'
                else:
                    order_item.payment_status = 'cancelled'

                order_item.product_variant.product.save()
                order_item.product_variant.save()

            if order.is_completed():
                order.complete_date = timezone.now()
                order.status = 'completed'
                order.save()

        order_item.save()
        return redirect(order_item_management, order.id)
    else:
        return redirect(admin_dashboard)
    

@login_required(login_url='admin_login')
@user_passes_test(is_admin, login_url='/permission-denied/')
def admin_coupon_management(request):
    coupons = Coupon.objects.all()
    search = request.GET.get('search')
    if search and search is not None:
        coupons = coupons.filter(Q(coupon_code__icontains=search) | Q(description__icontains=search))
        
    context = {
        'coupons' : coupons,
        'search' : search
    }
    return render(request, 'admin_page/coupon_management/admin_coupon_list.html', context)


@login_required(login_url='admin_login')
@user_passes_test(is_admin, login_url='/permission-denied/')
def add_coupon(request):
    if request.method == 'POST':
        coupon_code = request.POST.get('coupon_code')
        description = request.POST.get('description')
        minimum_amount = request.POST.get('minimum_amount')
        type = request.POST.get('type')
        discount = request.POST.get('discount')

        coupons = Coupon.objects.annotate(lower_name=Lower(Replace('coupon_code', Value(' '), Value('')))).values_list('lower_name', flat=True)
        coupon_check = (coupon_code.lower()).replace(" ", "")
        if coupon_check in coupons:
            messages.error(request, 'Coupon with similar Coupon Code exists')
            return redirect(add_coupon)

        try:
            coupon_code=(coupon_code.upper()).replace(" ", "")
            Coupon.objects.create(coupon_code=coupon_code, description=description, minimum_amount=minimum_amount, type = type, discount = discount)
            messages.success(request, 'Product Addedd Successfully')
        
        except IntegrityError:
            messages.error(request, 'Coupon with same Coupon Code exists')
            return redirect(admin_coupon_management)

        except:
            messages.error(request, 'Oops something went wrong')
            return redirect(admin_coupon_management)
        
        return redirect(admin_coupon_management)
    
    return render(request, 'admin_page/coupon_management/admin_add_coupon.html')

@login_required(login_url='admin_login')
@user_passes_test(is_admin, login_url='/permission-denied/')
def update_coupon(request, coupon_id):
    coupon = Coupon.objects.get(id=coupon_id)

    if request.method == 'POST':
        coupon_code = request.POST.get('coupon_code')
        description = request.POST.get('description')
        minimum_amount = request.POST.get('minimum_amount')
        type = request.POST.get('type')
        discount = request.POST.get('discount')
        is_expired = 'is_expired' in request.POST

        coupons = Coupon.objects.exclude(id=coupon_id).annotate(lower_name=Lower(Replace('coupon_code', Value(' '), Value('')))).values_list('lower_name', flat=True)
        coupon_check = (coupon_code.lower()).replace(" ", "")
        if coupon_check in coupons:
            messages.error(request, 'Coupon with similar Coupon Code exists')
            return redirect(update_coupon, coupon_id)

        try:
            change = False
            if coupon_code and coupon_code != coupon.coupon_code and coupon_code is not None:
                coupon_code=(coupon_code.upper()).replace(" ", "")
                coupon.coupon_code = coupon_code
                change = True
            
            if description and description != coupon.description and description is not None:
                coupon.description = description
                change = True

            if minimum_amount and float(minimum_amount) != coupon.minimum_amount and minimum_amount is not None:
                coupon.minimum_amount = minimum_amount
                change = True
            
            if type and type != coupon.type and type is not None:
                coupon.type = type
                change = True

            if discount and int(discount) != coupon.discount and discount is not None:
                coupon.discount = int(discount)
                change = True

            if is_expired != coupon.is_expired and is_expired is not None:
                coupon.is_expired = is_expired
                change = True
            
            if change:
                coupon.save()
                messages.success(request, 'Coupon Updated Successfully')
            else:
                messages.info(request,'No Changes detected')

            return redirect(admin_coupon_management)
        
        except:
            messages.error(request,'Oops... Something went wrong')
            return redirect(admin_coupon_management)
    
    context = {
        'coupon' : coupon,
    }
    return render(request, 'admin_page/coupon_management/update_coupon.html', context)


import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse

def render_to_excel(order_items, context):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ZKart Sales Report"

    cell = ws.cell(row=1, column=1, value="ZKart Sales Report")
    cell.font = Font(bold=True)
    ws.cell(row=2, column=1, value="Overall Order Items")
    ws.cell(row=2, column=2, value=context['overall_order_items'])

    ws.cell(row=3, column=1, value="Overall Order Amount")
    ws.cell(row=3, column=2, value=context['overall_order_amount'])

    ws.cell(row=4, column=1, value="Overall Order Discount")
    ws.cell(row=4, column=2, value=context['overall_order_discount'])

    headers = [
        "Order Id", "Username", "Order Item Id", "Order Date",
        "Product Title", "Product Original Price", "Sold Price",
        "Product Quantity", "Product Discount", "Coupon Applied",
        "Coupon Discount", "Total Amount", "Order Item Status",
        "Payment Status", "Payment Method"
    ]
    # Create header row
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num, value=header)
        cell.font = Font(bold=True)

    # Add data rows
    for row_num, order_item in enumerate(order_items, 7):
        ws.cell(row=row_num, column=1, value=order_item.order.order_identifier)
        ws.cell(row=row_num, column=2, value=order_item.order.customer.user.username)
        ws.cell(row=row_num, column=3, value=order_item.id)
        if order_item.order.order_date:
            order_date_naive = order_item.order.order_date.replace(tzinfo=None)
            order_date_str = order_date_naive.strftime("%Y-%m-%d")
        else:
            order_date_str = 'N/A'
        ws.cell(row=row_num, column=4, value=order_date_str)
        ws.cell(row=row_num, column=5, value=order_item.product_variant.product.title)
        ws.cell(row=row_num, column=6, value=order_item.original_price)
        ws.cell(row=row_num, column=7, value=order_item.selling_price)
        ws.cell(row=row_num, column=8, value=order_item.quantity)
        ws.cell(row=row_num, column=9, value=order_item.item_discount())
        ws.cell(row=row_num, column=10, value=order_item.order.coupon.coupon_code if order_item.order.coupon else 'N/A')
        ws.cell(row=row_num, column=11, value=order_item.coupon_discount)
        ws.cell(row=row_num, column=12, value=order_item.item_grand_total())
        ws.cell(row=row_num, column=13, value=order_item.status)
        ws.cell(row=row_num, column=14, value=order_item.payment_status)
        ws.cell(row=row_num, column=15, value=order_item.order.payment_method)


    # Save the workbook to a BytesIO object
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=zkart_sales_report.xlsx'
    wb.save(response)
    return response


from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO

def render_to_pdf(template_src, context_dict): 
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        response = HttpResponse(result.getvalue(), content_type='application/pdf')
        return response
    return None


@login_required(login_url='admin_login')
@user_passes_test(is_admin, login_url='/permission-denied/')
def sales_report(request):

    def grand_total(order_items):
        return sum(item.item_grand_total() for item in order_items)

    def grand_discount(order_items):
        return sum(item.item_grand_discount() for item in order_items)
    
    time_period = request.GET.get('time_period', 'all')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    single_date = request.GET.get('single_date')
    search = request.GET.get('search')

    pdf_button = request.GET.get('pdf_button')
    excel_button = request.GET.get('excel_button')

    order_items = OrderItem.objects.exclude(order__status = 'cart')
    
    if search and search is not None:
        order_items = order_items.filter(
                Q(order__order_identifier__icontains=search) |
                Q(order__customer__user__username__icontains=search) |
                Q(product_variant__product__title__icontains=search)
            )

    if time_period == 'today':
        order_items = order_items.filter(order__order_date__date=date.today())

    elif time_period == 'this_week':
        start_of_week = timezone.now() - timedelta(days=timezone.now().weekday())
        start_of_week = start_of_week.replace(hour=0, minute=0, second=0, microsecond=0)
        order_items = order_items.filter(order__order_date__gte=start_of_week)

    elif time_period == 'this_month':
        start_of_month = timezone.now().replace(day=1)
        order_items = order_items.filter(order__order_date__gte=start_of_month)

    elif time_period == 'this_year':
        start_of_year = timezone.now().replace(month=1, day=1)
        order_items = order_items.filter(order__order_date__gte=start_of_year)
    
    elif time_period == 'past_7_days':
        past_7_days = timezone.now() - timedelta(days=7)
        order_items = order_items.filter(order__order_date__gte=past_7_days)

    elif time_period == 'by_date' and single_date:
        order_items = order_items.filter(order__order_date__date=single_date)

    elif time_period == 'custom' and start_date and end_date:
        order_items = order_items.filter(order__order_date__date__range=[start_date, end_date])
    
    else:
        time_period = 'all'
        
    if time_period != 'custom':
        start_date = None
        end_date = None

    if time_period != 'by_date':
        single_date = None

    overall_order_items = order_items.count()
    overall_order_amount = grand_total(order_items)
    overall_order_discount = grand_discount(order_items)
    
    success_items = order_items.filter(status = 'delivered')
    overall_success_items = success_items.count()
    overall_success_amount = grand_total(success_items)
    overall_success_discount = grand_discount(success_items)

    cancelled_items = order_items.filter(status = 'cancelled')
    overall_cancelled_items = cancelled_items.count()
    overall_cancelled_amount = grand_total(cancelled_items)
    overall_cancelled_discount = grand_discount(cancelled_items)

    returned_items = order_items.filter(status = 'returned')
    overall_returned_items = returned_items.count()
    overall_returned_amount = grand_total(returned_items)
    overall_returned_discount = grand_discount(returned_items)

    requested_items = order_items.filter(status = 'return_request')
    overall_requested_items = requested_items.count()
    overall_requested_amount = grand_total(requested_items)
    overall_requested_discount = grand_discount(requested_items)

    in_progress_items = order_items.filter(Q(status = 'in_progress') | Q(status = 'shipped'))
    overall_in_progress_items = in_progress_items.count()
    overall_in_progress_amount = grand_total(in_progress_items)
    overall_in_progress_discount = grand_discount(in_progress_items)

    context = {
        'order_items' : order_items,
        'overall_order_items' : overall_order_items,
        'overall_order_amount' : overall_order_amount,
        'overall_order_discount' : overall_order_discount,

        'overall_success_items' : overall_success_items,
        'overall_success_amount' : overall_success_amount,
        'overall_success_discount' : overall_success_discount,

        'overall_cancelled_items' : overall_cancelled_items,
        'overall_cancelled_amount' : overall_cancelled_amount,
        'overall_cancelled_discount' : overall_cancelled_discount,

        'overall_returned_items' : overall_returned_items,
        'overall_returned_amount' : overall_returned_amount,
        'overall_returned_discount' : overall_returned_discount,

        'overall_requested_items' : overall_requested_items,
        'overall_requested_amount' : overall_requested_amount,
        'overall_requested_discount' : overall_requested_discount,

        'overall_in_progress_items' : overall_in_progress_items,
        'overall_in_progress_amount' : overall_in_progress_amount,
        'overall_in_progress_discount' : overall_in_progress_discount,

        'time_period' : time_period,
        'start_date' : start_date,
        'end_date' : end_date,
        'single_date' : single_date,
        'search' : search,
    }

    if pdf_button and pdf_button is not None:
        response = render_to_pdf('admin_page/order_management/pdf_sales_report.html', context)
        response['Content-Disposition'] = 'filename="sales_report.pdf"'
        return response

    if excel_button and excel_button is not None:
        return render_to_excel(order_items, context)

    return render(request, 'admin_page/order_management/sales_report.html', context)


# test
from geopy.geocoders import Nominatim
from geopy import distance
from django.views.decorators.csrf import csrf_exempt
from decimal import Decimal

def test_purpose(request):
    geolocator = Nominatim(user_agent="distance calculation")
    pin_code1, pin_code2 = '676105', 'Aliganj'
    coordinate1 = geolocator.geocode('676105')
    coordinate2 = geolocator.geocode('Aliganj')
    lat1, long1 = (coordinate1.latitude), (coordinate1.longitude)
    lat2, long2 = (coordinate2.latitude), (coordinate2.longitude)
    place1 = (11.075, 76.125) #(lat1, long1)
    place2 = (37.4219999, -122.0840575) #(lat2, long2)
    location = geolocator.reverse("37.4219999, -122.0840575")
    dec = '-123.23456'
    dec = Decimal(dec)
    places = coordinate2, coordinate1
    difference = distance.distance(place1, place2)
    context = {
        'distance': places
    }
    return render(request, 'test_purpose.html', context)

@csrf_exempt
def payment_success(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        try:
            client.utility.verify_payment_signature(data)
            return JsonResponse({'status': 'success'})
        except razorpay.errors.SignatureVerificationError:
            return JsonResponse({'status': 'failed'})

    return JsonResponse({'status': 'failed'})



